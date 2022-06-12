from get_data import data
from openpyxl.workbook import Workbook
from openpyxl.styles import Border, Side, Font, Alignment


class excel:
    def __init__(self):
        self.wb = Workbook()
        ws1 = self.wb.active
        ws1.title = "Analysis"
        self.into_excel()
        del self.wb['Analysis']
        self.wb.save("CTU FIT Bachelor Specializations.xlsx")

    def into_excel(self):
        unformated_data = list(
            data('https://fit.cvut.cz/en/studies/programs-and-specializations/bachelor'))
        for i in range(len(unformated_data)):
            sposition = 4
            if unformated_data[i][0] == "Software Engineering":
                ws = self.wb.create_sheet(f"{unformated_data[i][0]}", 0)
            else:
                ws = self.wb.create_sheet(f"{unformated_data[i][0][:31]}")
            ws.merge_cells("A1:D3")
            header = ws['A1']
            header.font = Font(
                name='Cavolini', color='0033cc', size=30, bold=True)
            _ = ws.cell(column=1, row=1, value=f"{unformated_data[i][0]}")
            colors = ["99ff66", "3399ff", "ff3399",
                      "ff9933", "cc33ff", "66ffff"]
            for item in unformated_data[i][1]:
                _ = ws.cell(column=1, row=sposition,
                            value=f"{item[0]}")  # semester
                ws.merge_cells(start_row=sposition, start_column=1,
                               end_row=1+sposition, end_column=4)
                semesterheader = ws.cell(column=1, row=sposition)
                semesterheader.font = Font(
                    name='Times New Roman', color=colors[-1], size=20)
                colors.pop()
                semesterheader.alignment = Alignment(horizontal="right")
                my_cell = ws.cell(column=1, row=sposition +
                                  2, value="Code")  # header
                my_cell.font = Font(name='New Times Roman',
                                    color="000000", size=16, bold=True)
                my_cell.border = Border(top=Side(border_style='thin', color="000000"),
                                        left=Side(border_style='thin',
                                                  color="000000"),
                                        bottom=Side(
                                            border_style='thin', color="000000"),
                                        right=Side(border_style='thin', color="000000"))
                my_cell.alignment = Alignment(horizontal='center')
                my_cell = ws.cell(column=2, row=sposition+2,
                                  value="Subject")  # header
                my_cell.font = Font(name='New Times Roman',
                                    color="000000", size=16, bold=True)
                my_cell.border = Border(top=Side(border_style='thin', color="000000"),
                                        left=Side(border_style='thin',
                                                  color="000000"),
                                        bottom=Side(
                                            border_style='thin', color="000000"),
                                        right=Side(border_style='thin', color="000000"))
                my_cell.alignment = Alignment(horizontal='center')
                my_cell = ws.cell(column=3, row=sposition+2,
                                  value="Credits")  # header
                my_cell.font = Font(name='New Times Roman',
                                    color="000000", size=16, bold=True)
                my_cell.border = Border(top=Side(border_style='thin', color="000000"),
                                        left=Side(border_style='thin',
                                                  color="000000"),
                                        bottom=Side(
                                            border_style='thin', color="000000"),
                                        right=Side(border_style='thin', color="000000"))
                my_cell.alignment = Alignment(horizontal='center')
                my_cell = ws.cell(column=4, row=sposition +
                                  2, value="Link")  # header
                my_cell.font = Font(name='New Times Roman',
                                    color="000000", size=16, bold=True)
                my_cell.border = Border(top=Side(border_style='thin', color="000000"),
                                        left=Side(border_style='thin',
                                                  color="000000"),
                                        bottom=Side(
                                            border_style='thin', color="000000"),
                                        right=Side(border_style='thin', color="000000"))
                my_cell.alignment = Alignment(horizontal='center')
                # print(item[0]) # semester
                increment = 1
                for sub_item in item[1]:
                    if sub_item != []:
                        my_cell = ws.cell(
                            column=1, row=sposition+2+increment, value=f"{sub_item[0]}")  # data
                        my_cell.font = Font(
                            name='Ariel', color="000000", size=12)
                        my_cell.border = Border(top=Side(border_style='thin', color="000000"),
                                                left=Side(
                                                    border_style='thin', color="000000"),
                                                bottom=Side(
                                                    border_style='thin', color="000000"),
                                                right=Side(border_style='thin', color="000000"))
                        my_cell.alignment = Alignment(horizontal='center')
                        my_cell = ws.cell(
                            column=2, row=sposition+2+increment, value=f"{sub_item[1][1]}")  # data
                        my_cell.font = Font(
                            name='Ariel', color="000000", size=12)
                        my_cell.border = Border(top=Side(border_style='thin', color="000000"),
                                                left=Side(
                                                    border_style='thin', color="000000"),
                                                bottom=Side(
                                                    border_style='thin', color="000000"),
                                                right=Side(border_style='thin', color="000000"))
                        my_cell.alignment = Alignment(horizontal='center')
                        try:
                            value = int(sub_item[-1])
                        except:
                            value = ""
                        my_cell = ws.cell(
                            column=3, row=sposition+2+increment, value=value)  # data
                        my_cell.font = Font(
                            name='Ariel', color="000000", size=12)
                        my_cell.border = Border(top=Side(border_style='thin', color="000000"),
                                                left=Side(
                                                    border_style='thin', color="000000"),
                                                bottom=Side(
                                                    border_style='thin', color="000000"),
                                                right=Side(border_style='thin', color="000000"))
                        my_cell.alignment = Alignment(horizontal='center')
                        my_cell = ws.cell(
                            column=4, row=sposition+2+increment, value=f"{sub_item[1][0]}")  # data
                        my_cell.font = Font(
                            name='Ariel', color="000000", size=12)
                        my_cell.border = Border(top=Side(border_style='thin', color="000000"),
                                                left=Side(
                                                    border_style='thin', color="000000"),
                                                bottom=Side(
                                                    border_style='thin', color="000000"),
                                                right=Side(border_style='thin', color="000000"))
                        my_cell.alignment = Alignment(horizontal='center')
                        increment += 1
                        # print(sub_item) # row of data for the given semester
                # this is to position the semester to the right row
                sposition += len(item[1])+3
            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 30
            ws.column_dimensions["C"].width = 15
            ws.column_dimensions["D"].width = 50
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
